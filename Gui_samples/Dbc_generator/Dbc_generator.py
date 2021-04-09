"""
Author : Shriram Sundaram
Date : 08.04.2021
Description : Generation of CAN DBC file from User uploaded Excel Sheet.
This application includes front end GUI which is created using pyQt5
and backend process using openpyxl and basic file operation modules.

"""

from openpyxl import load_workbook
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMainWindow, QFileDialog
import os


class DBCGenerator(QMainWindow):
    def __init__(self):
        """
        DBCGenerator is a class, which inherits another class called QMainWindow.
        DBCGenerator has following methods:
                -- __Init__ : This method basically initializes the UiDesign with basic set of parameters.
                Main function is to set up the QMainWindow with the Size,Color,Shape
                and Adding Buttons,Display Text on to the screen
                -- translate_ui :
                -- load_workbook :
                -- excel_extract_data:
                -- dbc_sketch:
                -- dbc_write_frames_signal:
                -- clear_data:
        """
        self.col = 0
        self.rows = 5
        self.CAN_frame_data = {}
        self.CAN_signal_data = {}
        self.open_excel = ""
        self.sheet = ""
        self.header_list = ['CAN_ID', 'CAN Message', 'DLC', 'CAN Signal', 'Receiver Node', 'Transmitter Node',
                            'Signal Start Bit',
                            'Signal Byte Order', 'Signal Value Type', 'Signal Factor', 'Signal offset', 'Signal Min',
                            'Signal Max']
        self.dbc_file = ""
        self.dbc_obj = ""
        self.dbc_info = ""
        self.var = ""

        super(DBCGenerator, self).__init__()
        self.setObjectName("DBCGenerator")
        self.resize(637, 382)
        self.central_widget = QtWidgets.QWidget(self)
        self.central_widget.setObjectName("central_widget")

        self.line_edit = QtWidgets.QLineEdit(self.central_widget)
        self.line_edit.setGeometry(QtCore.QRect(70, 110, 351, 22))
        self.line_edit.setObjectName("line_edit")

        self.line_edit_row_size = QtWidgets.QLineEdit(self.central_widget)
        self.line_edit_row_size.setGeometry(QtCore.QRect(70, 160, 51, 22))
        self.line_edit_row_size.setObjectName("line_edit_row_size")

        self.generate_button = QtWidgets.QPushButton(self.central_widget)
        self.generate_button.setGeometry(QtCore.QRect(180, 190, 93, 28))
        self.generate_button.setObjectName("generate_button")

        self.browse_button = QtWidgets.QPushButton(self.central_widget)
        self.browse_button.setGeometry(QtCore.QRect(420, 110, 51, 22))
        self.browse_button.setObjectName("browse_button")

        self.label = QtWidgets.QLabel(self.central_widget)
        self.label.setGeometry(QtCore.QRect(70, 65, 121, 31))
        self.label_log = QtWidgets.QLabel(self.central_widget)
        self.label_log.setGeometry(QtCore.QRect(80, 260, 481, 31))
        font = QtGui.QFont()
        font.setFamily("Microsoft Himalaya")
        font.setPointSize(20)
        self.label_log.setFont(font)
        font.setUnderline(True)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_log.setObjectName("label_log")
        self.label_row_size = QtWidgets.QLabel(self.central_widget)
        self.label_row_size.setGeometry(QtCore.QRect(70, 140, 55, 16))
        self.label_row_size.setObjectName("label_row_size")
        self.setCentralWidget(self.central_widget)

        self.translate_ui()
        self.browse_button.clicked.connect(self.load_workbook)

    def translate_ui(self):
        """
        sets name for all Buttons, line_edits(input from user),label (output display: eg:row size)
        :param :instance is created after Init
        :return:
        """
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("DBCGenerator", "DBCGenerator"))
        self.generate_button.setText(_translate("DBCGenerator", "Generate"))
        self.browse_button.setText(_translate("DBCGenerator", "Browse"))
        self.label.setText(_translate("DBCGenerator", "Import DBC"))
        self.label_log.setText(_translate("MainWindow", "Log: Enter Row Size"))
        self.label_row_size.setText(_translate("MainWindow", "Row size"))

    def load_workbook(self):
        """
        load pop up window(goes to Directory and choose the File(.xlsx)) after browse button is pressed
        :param :instance is created after browse_button is pressed
        :return:
        """
        open_file = (QFileDialog.getOpenFileName(None, "Select a File", os.path.expanduser("~")))
        print((open_file[0]))
        if open_file != ("", ""):  # User can press cancel or x button. QFileDialog.getOpenFileName returns tuple()
            self.line_edit.setText(str(open_file[0]))
            self.open_excel = load_workbook(filename=str(open_file[0]))
            self.sheet = self.open_excel['CAN_Sheet']
            self.generate_button.clicked.connect(self.excel_extract_data)
        else:
            self.label_log.setText("Log: Input file is not loaded")

    def excel_extract_data(self):
        """
        This function extracts the data from loaded Excel File(.xlsx).
        :param :instance is created after generate_button is pressed
        :return:
        """
        if len(self.line_edit_row_size.text()) != 0:
            self.rows = int(self.line_edit_row_size.text())
            print("rows:", int(self.line_edit_row_size.text()))
        for self.row in self.sheet.iter_rows(min_row=1, max_col=self.col, max_row=self.rows, values_only=True):
            if self.row[self.col] not in self.header_list:
                if self.row[self.col] not in self.CAN_frame_data:
                    self.CAN_frame_data.update({self.row[self.col]: [self.row[self.col + 1],
                                                                     [self.row[self.col + 2], self.row[self.col + 5]]]})

        for row in self.sheet.iter_rows(min_row=1, max_col=self.col, max_row=self.rows, values_only=True):
            if row[self.col] not in self.header_list:
                if row[self.col] in self.CAN_frame_data:
                    self.CAN_signal_data.update({row[self.col + 3]: [row[self.col],
                                                                     row[self.col + 6], row[self.col + 7],
                                                                     row[self.col + 8], row[self.col + 9],
                                                                     row[self.col + 10], row[self.col + 11],
                                                                     row[self.col + 12], row[self.col + 13],
                                                                     row[self.col + 14], row[self.col + 4]]})

        print(self.CAN_frame_data)
        print(self.CAN_signal_data)
        self.dbc_sketch()

    def dbc_sketch(self):
        """
        This function writes basic data to the DBC file. (operations: file open and file write)
        :param :instance is created after excel_extract_data() is processed
        :return:
        """
        self.dbc_file = "generated_DBC.dbc"
        self.dbc_info = "VERSION " + "  \n"
        self.var = ("NS_:\n" +
                    "   NS_DESC_\n" + "   CM_\n" + "   BA_DEF_\n" + "   BA_\n" + "   VAL_\n" + "   CAT_DEF_\n" + "   CAT_\n"
                    + "   FILTER\n" + "   BA_DEF_DEF_\n" + "   EV_DATA_\n"
                    + "   ENVVAR_DATA_\n" + "   SGTYPE_\n" + "   SGTYPE_VAL_\n" + "   BA_DEF_SGTYPE_\n" + "   BA_SGTYPE_\n" + "   SIG_TYPE_REF_\n"
                    + "   VAL_TABLE_\n" + "   SIG_GROUP_\n" + "   SIG_VALTYPE_\n" + "   SIGTYPE_VALTYPE_\n" + "   BO_TX_BU_\n" + "   BA_DEF_REL_\n"
                    + "   BA_REL_\n" + "   BA_DEF_DEF_REL_\n" + "   BU_SG_REL_\n" + "   BU_EV_REL_\n" + "   BU_BO_REL_\n " + "  SG_MUL_VAL_\n" +
                    "BS_:\n")

        self.dbc_obj = open(self.dbc_file, "w")
        self.dbc_obj.writelines(self.dbc_info + self.var)
        self.dbc_obj.close()
        self.dbc_write_frames_signal()

    def dbc_write_frames_signal(self):
        """
        This function writes important data(like CAN Frame and Signal information respectively) to the DBC file.
        (operations: writes data from CAN_frame_data() and CAN_signal_data() to DBC file)
        :param :instance is created after dbc_sketch() is processed
        :return:
        """
        self.dbc_obj = open(self.dbc_file, "a")
        BU_ = False
        ByteOrder = "@1"
        uInt = "+"
        for value in self.CAN_frame_data:
            print(
                "\n" + "BO_ " + str(value) + " " + str(self.CAN_frame_data[value][0]) + ":" + " " + str(
                    self.CAN_frame_data[value][1][0]) + " " + str(self.CAN_frame_data[value][1][1]))
            new_frame_id = True
            if not BU_:
                self.dbc_obj.writelines("\n" + "BU_ " + str(self.CAN_frame_data[value][1][1]) + "\n" + "\n")
                BU_ = True
                new_frame_id = False
            if new_frame_id:
                self.dbc_obj.writelines("\n")
            self.dbc_obj.writelines(
                "\n" + "BO_ " + str(value) + " " + str(self.CAN_frame_data[value][0]) + ":" + " " + str(
                    self.CAN_frame_data[value][1][0]) + " " + str(self.CAN_frame_data[value][1][1]))
            for signalName in self.CAN_signal_data:
                print(self.CAN_signal_data[signalName])
                if self.CAN_signal_data[signalName][0] == value:
                    if self.CAN_signal_data[signalName][3] != "Intel":
                        ByteOrder = "@0"
                    if self.CAN_signal_data[signalName][4] != "unsigned Integer":
                        uInt = "-"
                    self.dbc_obj.writelines(
                        "\n" + " SG_ " + str(signalName) + " : " + str(self.CAN_signal_data[signalName][1]) + "|" +
                        str(self.CAN_signal_data[signalName][2]) + ByteOrder + uInt + " (" + str(
                            self.CAN_signal_data[signalName][5])
                        + "," + str(self.CAN_signal_data[signalName][6]) + ")" + " [" + str(
                            self.CAN_signal_data[signalName][7]) + "|"
                        + str(self.CAN_signal_data[signalName][8]) + "]" + " " + '"' + str(
                            self.CAN_signal_data[signalName][9])
                        + '"' + " " + str(self.CAN_signal_data[signalName][10]))
        self.dbc_obj.close()
        self.clear_data()

    def clear_data(self):
        """
        This function clears all important data  from CAN_frame_data() and CAN_signal_data().
         (operations: writes data from CAN_frame_data() and CAN_signal_data() to DBC file)
        :param :instance is created after dbc_write_frames_signal() is processed
        :return:
        """
        self.label_log.setText("Log: DBC file generated Successfully")
        self.CAN_signal_data.clear()
        self.CAN_frame_data.clear()


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    ui = DBCGenerator()
    ui.show()
    sys.exit(app.exec_())
